import os
import csv
import requests as re
import openpyxl as opxl
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException



# initial setup
url = 'https://www.ebay.com/'
item_to_search = input('INSERT ITEM TO SEARCH\n')             # define your search here

driver = webdriver.Firefox()
driver.get(url)

# find search-bar element. enter item to search if found
search_bar = driver.find_element_by_xpath("//input[@id='gh-ac']")      # find search-bar element
if search_bar:
    search_bar.click()
    search_bar.send_keys(item_to_search)
    search_bar.send_keys(Keys.RETURN)
else:
    print('no search-bar element found')
    driver.quit()

# listing types: wait for elements to load before searching for them
try:
    buy_it_now = WebDriverWait(driver, 8).until(EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div[4]/div[6]/div[1]/div/div[1]/div[2]/div/div/ul/li[4]/a")))
except TimeoutException:
    print('loading took too much time')
    driver.quit()
    exit()
except NoSuchElementException:
    print('button not found')
    driver.quit()
    exit()

buy_it_now.click()

# add 'free international shipping' filter to search:
free_shipping_checkbox = WebDriverWait(driver, 3).until(EC.element_to_be_clickable(
    (By.XPATH, "/html/body/div[4]/div[5]/ul/li[1]/ul/li[6]/div[2]/ul/li/div/a/div/div/span[1]")))
if free_shipping_checkbox:
    free_shipping_checkbox.click()

# grab current page's url and open it with BeautifulSoup lxml parser
req = re.get(driver.current_url)
driver.quit()                             # exit browser, not relevant anymore
soup = BeautifulSoup(req.content, 'lxml')

# get all items from website
items = soup.find_all('div', {'class': 's-item__info clearfix'})

items_list = []
# fill items_list with dicts of name, price, link
for item in items:
    # if item has price-range, retrieve highest, else retrieve the price, and convert to float
    price = item.find('span', {'class': 's-item__price'}).get_text().split()[-1]

    # remove comma from price
    price = price.replace(',', '')

    # fill info
    item_info = {
    'name' : item.find('h3').get_text(),
    'price' : float(price),
    'link' : item.find('a', href=True)['href']
    }
    items_list.append(item_info)

#sort list by price: high to low
items_list.sort(key=lambda x: x['price'], reverse=False)

#check if list is empty
if not items_list:
    print("No items found")
# print each item's info
for item in items_list:
    print(f"name = {item['name']}")
    print(f"price = {item['price']}")
    print(f"link = {item['link']}", end="\n\n")

# into CSV
number_of_sheets = 2                                    # define number of wanted sheets
items_in_sheet = (len(items_list)//number_of_sheets)
j = 0
for i in range(1, number_of_sheets+1):                  # each iteration creates a new csv file ('sheet')
    with open(f'sheet{i}.csv', 'wt') as f:
        f_writer = csv.writer(f)
        while j < (items_in_sheet * i):
            try:
                f_writer.writerow([items_list[j]['name'], items_list[j]['price'], items_list[j]['link']])
            except UnicodeEncodeError:
                pass
            j += 1

# merge CSV files into a single file with multiple sheets:
workbook = opxl.Workbook()
workbook.active.title = 'page1'           # workbook start with 1 sheet by default, name it page1

# create other sheets starting with name page2
for i in range(1, number_of_sheets):
    workbook.create_sheet(f"page{i+1}")

# for each sheet, fill one CVS file info
for i in range(1, number_of_sheets+1):
    with open(f'sheet{i}.csv', 'r') as f:        # open i'th csv file in read mode
        f_reader = csv.reader(f)                 # create reader object for current file
        for row in f_reader:                     # for each row in csv file, append into current sheet
            workbook[f"page{i}"].append(row)

# delete csv files after merge (starting with 2nd sheet since first is created by default in workbook):
for i in range(1, number_of_sheets+1):
    os.remove(f'sheet{i}.csv')

# save final merged file
workbook.save('ebay_scraper.xlsx')

# open merged file
workbook = load_workbook('ebay_scraper.xlsx')
if workbook:
    # fit cell's size to text
    for sheet in workbook.worksheets:                            # loop through all sheets in file
        for col in sheet.columns:                     # loop through all columns in file
            max_length = 0
            column_name = col[0].column_letter             # Get the column name(letter to be used as key)

            # loop through all cells in column, and find largest text string's size
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            # change cell's width
            sheet.column_dimensions[column_name].width = (max_length + 2) * 1.02      # change column's width
    workbook.save('ebay_scraper.xlsx')
else:
    print('ERROR loading xlsx file.')

